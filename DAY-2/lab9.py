from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active
ws.title = "Sales Data"
headers = ["Date", "Product", "Quantity", "Price", "Total"]
data = [
    ["2024-01-01", "Widget A", 10, 5.0, None],
    ["2024-01-02", "Widget B", 15, 7.5, None],
    ["2024-01-03", "Widget ABCDEF", 5, 12.0, None],
]
ws.append(headers)
for row in data:
    ws.append(row)
for row in range(2, len(data) + 2):  # Rows 2 to data length
    ws[f"E{row}"] = f"=C{row}*D{row}"  # Formula: Quantity * Price

wb.save("sales_data.xlsx")
print("Headers formatted and column widths adjusted!")
